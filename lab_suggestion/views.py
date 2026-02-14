from django.shortcuts import render, redirect
from .models import Test, Lab, ContactMessage, ChatMessage, AIRecommendation, TestBooking
from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import get_object_or_404
from .forms import LabUserRegistrationForm, TestForm, ContactForm, LabForm, ExcelUploadForm, AdminLabEditForm, TestBookingForm
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.views import LoginView
import openpyxl
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.forms import modelformset_factory # Import modelformset_factory
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import uuid
from .ai_service import AIChatbotService, AIRecommendationService
from .email_utils import send_booking_confirmation_email, send_booking_update_email, send_booking_cancellation_email

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

# Create your views here.
def index(request):
    # Get popular tests - prioritize tests with prices, limit to 8 for better display
    # Filter out obviously invalid test names
    valid_tests = Test.objects.exclude(
        name__in=['a', 'u', 'z', 'x', 'acer', 'xray', 'ICU', 'BED', 'SSCU',
                  'AMBULANCE CHARG', 'VENTILATOR CHARGE', 'CABIN BED', 'OBSERVATION BED',
                  'POST ANESTHESIA BED', 'TRANSPLANT ROOM', 'TRIPLE BED']
    ).filter(price__isnull=False).order_by('price')[:8]
    
    if valid_tests.count() < 4:
        # Fallback to any valid tests if not enough with prices
        valid_tests = Test.objects.exclude(
            name__in=['a', 'u', 'z', 'x', 'acer', 'xray', 'ICU', 'BED', 'SSCU',
                      'AMBULANCE CHARG', 'VENTILATOR CHARGE', 'CABIN BED', 'OBSERVATION BED',
                      'POST ANESTHESIA BED', 'TRANSPLANT ROOM', 'TRIPLE BED']
        )[:8]
    
    return render(request, 'index.html', {'popular_tests': valid_tests})


@require_http_methods(["GET"])
def search_tests_autocomplete(request):
    """API endpoint for test autocomplete search"""
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 1:
        return JsonResponse({'results': []})
    
    # Define junk test names to exclude
    junk_tests = ['a', 'u', 'z', 'x', 'acer', 'xray', 'ICU', 'BED', 'SSCU',
                  'AMBULANCE CHARG', 'VENTILATOR CHARGE', 'CABIN BED', 'OBSERVATION BED',
                  'POST ANESTHESIA BED', 'TRANSPLANT ROOM', 'TRIPLE BED', 'VENTIL']
    
    # Search for tests matching the query (case-insensitive)
    # Filter out junk tests
    tests = Test.objects.filter(
        name__icontains=query
    ).exclude(
        name__in=junk_tests
    )[:20]
    
    # Remove duplicates by using a dictionary with test name as key
    seen_names = {}
    for test in tests:
        if test.name not in seen_names:
            seen_names[test.name] = {
                'id': test.id,
                'name': test.name,
                'price': float(test.price) if test.price else None,
                'description': test.description
            }
    
    results = list(seen_names.values())
    return JsonResponse({'results': results})

def lab_login_view(request):
    """Custom login view for lab users"""
    if request.user.is_authenticated:
        # If user is already logged in, redirect based on user type
        if request.user.is_superuser:
            return redirect('admin_lab_list')
        elif hasattr(request.user, 'lab'):
            return redirect('manage_lab')
        else:
            return redirect('index')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect based on user type
                if user.is_superuser:
                    return redirect('admin_lab_list')
                elif hasattr(user, 'lab'):
                    return redirect('manage_lab')
                else:
                    return redirect('index')
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})

def admin_login_view(request):
    """Admin login view - only allows superuser access"""
    if request.user.is_authenticated and request.user.is_superuser:
        # If already logged in as superuser, redirect to admin panel
        return redirect('admin_lab_list')
    
    error_message = None
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        if not username or not password:
            error_message = 'Please enter both username and password.'
        else:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_superuser:
                    login(request, user)
                    messages.success(request, f'Welcome back, {user.username}!')
                    return redirect('admin_lab_list')
                else:
                    error_message = 'Access denied. This account does not have administrator privileges.'
            else:
                error_message = 'Invalid username or password. Please try again.'
    
    return render(request, 'adminlogin.html', {'error': error_message})


def about_page(request):
    return render(request, 'about.html')

def contact_page(request):
    # Check if this is a booking request
    is_booking = request.GET.get('booking') == 'true'
    test_name = request.GET.get('test_name', '')
    lab_name = request.GET.get('lab_name', '')
    
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            recipient_choice = form.cleaned_data['recipient_choice']
            contact_message = form.save(commit=False)
            if recipient_choice == 'admin':
                contact_message.recipient_admin = True
                contact_message.lab = None
            else:
                contact_message.recipient_admin = False
                contact_message.lab = Lab.objects.get(id=int(recipient_choice))
            contact_message.save()
            if is_booking:
                messages.success(request, f'Your booking request for {test_name} has been sent to {lab_name}. They will contact you shortly!')
            else:
                messages.success(request, 'Your message has been sent successfully!')
            return redirect('index')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-fill form if it's a booking request
        initial_data = {}
        if is_booking and test_name and lab_name:
            # Find the lab by name
            try:
                lab = Lab.objects.get(name=lab_name)
                # Set the recipient to the lab
                initial_data['recipient_choice'] = str(lab.id)
                # Don't pre-fill message - let user add their own notes
                initial_data['message'] = ''
            except Lab.DoesNotExist:
                pass
        
        form = ContactForm(initial=initial_data)
    
    return render(request, 'contact.html', {
        'form': form, 
        'is_booking': is_booking,
        'test_name': test_name,
        'lab_name': lab_name
    })

def search_labs(request):
    query = request.GET.get('query')
    print(f"Search query received: {query}")
    display_results = [] # This will store a list of dictionaries, each representing a test-lab pair

    if query:
        # Define junk test names to exclude
        junk_tests = ['a', 'u', 'z', 'x', 'acer', 'xray', 'ICU', 'BED', 'SSCU',
                      'AMBULANCE CHARG', 'VENTILATOR CHARGE', 'CABIN BED', 'OBSERVATION BED',
                      'POST ANESTHESIA BED', 'TRANSPLANT ROOM', 'TRIPLE BED', 'VENTIL']
        
        # Find tests that match the query (case-insensitive)
        # Filter out junk tests
        matching_tests = Test.objects.filter(
            name__icontains=query
        ).exclude(
            name__in=junk_tests
        )
        print(f"Matching tests found: {matching_tests}")

        for test in matching_tests:
            # For each test, find all labs that offer it
            labs_offering_test = test.lab_set.all()
            print(f"Labs offering test '{test.name}': {labs_offering_test}")

            for lab in labs_offering_test:
                display_results.append({
                    'test_id': test.id,
                    'test_name': test.name,
                    'test_description': test.description,
                    'test_price': test.price,
                    'lab_id': lab.id,
                    'lab_name': lab.name,
                    'lab_address': lab.address,
                    'lab_contact_email': lab.contact_email,
                    'lab_contact_phone': lab.contact_phone,
                })

    return render(request, 'labdetails.html', {'display_results': display_results, 'query': query})

def lab_registration(request):
    if request.method == 'POST':
        form = LabUserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Create lab with default values for required fields
            Lab.objects.create(
                user=user,
                name=form.cleaned_data['name'],
                address=form.cleaned_data['address'],
                city='Kathmandu',  # Default city
                state='Bagmati',   # Default state
                zip_code='44600',  # Default zip code
                phone_number=form.cleaned_data['contact_phone'],
                contact_email=form.cleaned_data['contact_email'],
                contact_phone=form.cleaned_data['contact_phone']
            )
            messages.success(request, 'Registration successful! Please login to access your lab dashboard.')
            return redirect('login')  # Redirect to login page after successful registration
    else:
        form = LabUserRegistrationForm()
    return render(request, 'labregister.html', {'form': form})


@login_required
def manage_lab(request):
    try:
        # Get the lab associated with the logged-in user
        lab = request.user.lab
    except Lab.DoesNotExist:
        # If the user is a superuser without a lab, maybe show a different page or redirect
        if request.user.is_superuser:
            return redirect('admin_lab_list')
        # For regular users not associated with a lab, redirect to the landing page
        return redirect('index')

    if request.method == 'POST':
        form = TestForm(request.POST)
        if form.is_valid():
            # Create a new test and associate it with the current lab
            test = form.save()
            lab.tests.add(test)
            return redirect('manage_lab')
    else:
        form = TestForm()

    tests = lab.tests.all()
    # Fetch messages sent to this lab (recipient_admin is False)
    lab_messages = ContactMessage.objects.filter(lab=lab, recipient_admin=False).order_by('-sent_at')
    
    # Statistics for dashboard
    total_tests = tests.count()
    total_messages = lab_messages.count()
    recent_messages = lab_messages[:5]
    
    # Add ExcelUploadForm to context for lab-only upload interface
    excel_upload_form = ExcelUploadForm()
    
    context = {
        'lab': lab,
        'tests': tests,
        'form': form,
        'excel_upload_form': excel_upload_form,
        'lab_messages': lab_messages,
        'total_tests': total_tests,
        'total_messages': total_messages,
        'recent_messages': recent_messages,
    }
    return render(request, 'labmanage.html', context)


@login_required
def edit_test(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    # Check if the user is an admin or owns the lab that offers this test
    if not (request.user.is_superuser or (hasattr(request.user, 'lab') and test in request.user.lab.tests.all())):
        return redirect('manage_lab')

    if request.method == 'POST':
        form = TestForm(request.POST, instance=test)
        if form.is_valid():
            form.save()
            return redirect('manage_lab')
    else:
        form = TestForm(instance=test)

    return render(request, 'edit_test.html', {'form': form, 'test': test})


@login_required
def delete_test(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    lab = request.user.lab
    # Check if the user is an admin or owns the lab that offers this test
    if not (request.user.is_superuser or (hasattr(request.user, 'lab') and test in lab.tests.all())):
        return redirect('manage_lab')

    if request.method == 'POST':
        # Remove the test from the lab. It doesn't delete the test itself
        # in case other labs offer it.
        lab.tests.remove(test)
        return redirect('manage_lab')

    return render(request, 'delete_test_confirm.html', {'test': test})


@user_passes_test(lambda u: u.is_superuser)
def admin_lab_list(request):
    labs = Lab.objects.all()
    
    # Statistics for dashboard
    total_labs = labs.count()
    total_tests = Test.objects.count()
    total_messages = ContactMessage.objects.count()
    recent_messages = ContactMessage.objects.order_by('-sent_at')[:5]
    total_users = User.objects.filter(is_superuser=False).count()
    
    # Recent labs (last 5 registered)
    recent_labs = labs.order_by('-id')[:5]
    
    context = {
        'labs': labs,
        'total_labs': total_labs,
        'total_tests': total_tests,
        'total_messages': total_messages,
        'total_users': total_users,
        'recent_messages': recent_messages,
        'recent_labs': recent_labs,
    }
    return render(request, 'admin_lab_manage.html', context)

@user_passes_test(lambda u: u.is_superuser)
def admin_edit_lab(request, lab_id):
    lab = get_object_or_404(Lab, pk=lab_id)
    user = lab.user
    
    # Removed 'popularity' from fields list in modelformset_factory
    TestFormSet = modelformset_factory(Test, fields=('name', 'description', 'price'), extra=0, can_delete=True)

    if request.method == 'POST':
        print("Raw POST data:", request.POST) # Add this line
        # Bind form to existing lab instance and pass initial data for user-related fields only
        form = AdminLabEditForm(request.POST, instance=lab, initial={
            'username': user.username,
            'email': user.email,
        })
        test_form = TestForm(request.POST, prefix='test') # Initialize with prefix in POST request
        formset = TestFormSet(request.POST, queryset=lab.tests.all(), prefix='tests')

        if form.is_valid() and formset.is_valid():
            # Update User model fields
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.save()
            
            # Update Lab model fields
            lab.name = form.cleaned_data['name']
            lab.address = form.cleaned_data['address']
            lab.contact_email = form.cleaned_data['contact_email']
            lab.contact_phone = form.cleaned_data['contact_phone']
            lab.save()

            # Save changes to existing tests and handle deletions
            formset.save()
            for test_instance in formset.deleted_objects:
                lab.tests.remove(test_instance) # Disassociate deleted tests from the lab

            # Handle adding a new test ONLY if the name field is provided
            if request.POST.get('test-name') and test_form.is_valid(): # Check for 'test-name' with the correct prefix
                new_test = test_form.save()
                lab.tests.add(new_test)
                messages.success(request, 'New test added and associated with the lab.')

            messages.success(request, 'Lab and associated tests updated successfully.')
            return redirect('admin_lab_list')
        else:
            print("Form errors:", form.errors) # Add this line to inspect form errors
            print("Test Form errors:", test_form.errors) # Add this line to inspect new test form errors
            print("Formset errors:", formset.errors) # Add this line to inspect formset errors
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AdminLabEditForm(instance=lab, initial={
            'username': user.username,
            'email': user.email,
        })
        test_form = TestForm(prefix='test')  # Add prefix to avoid field name conflict
        formset = TestFormSet(queryset=lab.tests.all(), prefix='tests')

    # Get all available tests for selection (optional, depending on desired UI for adding existing tests)
    all_tests = Test.objects.all()

    return render(request, 'admin_edit_lab.html', {
        'form': form,
        'lab': lab,
        'test_form': test_form, # Form for adding new tests
        'formset': formset,     # Formset for managing existing tests
        'all_tests': all_tests  # All tests for potential selection
    })

@user_passes_test(lambda u: u.is_superuser)
def admin_delete_lab(request, lab_id):
    lab = Lab.objects.get(id=lab_id)
    lab.user.delete() # This will also delete the lab due to the CASCADE on the OneToOneField
    return redirect('admin_lab_list')

@user_passes_test(lambda u: u.is_superuser)
def view_contacts(request):
    messages = ContactMessage.objects.all().order_by('-sent_at')
    return render(request, 'view_contacts.html', {'messages': messages})


@user_passes_test(lambda u: u.is_superuser)
def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                # Assuming the first row is headers
                headers = [cell.value.lower() if cell.value else None for cell in sheet[1]]

                # Basic validation for required columns
                required_lab_cols = ['lab name', 'address', 'city', 'state', 'zip code', 'phone number', 'contact email', 'contact phone']
                required_test_cols = ['test name', 'test description', 'test price']

                if not all(col in headers for col in required_lab_cols + required_test_cols):
                    messages.error(request, "Missing one or more required columns in the Excel file (case-insensitive). Required: Lab Name, Address, City, State, Zip Code, Phone Number, Contact Email, Contact Phone, Test Name, Test Description, Test Price.")
                    return render(request, 'admin_upload_excel.html', {'form': form})

                for row_index in range(2, sheet.max_row + 1):
                    row_data = {headers[i]: cell.value for i, cell in enumerate(sheet[row_index])}

                    # Create or get Lab
                    lab_name = row_data.get('lab name')
                    if not lab_name:
                        messages.warning(request, f"Skipping row {row_index}: Lab Name is missing.")
                        continue

                    try:
                        lab = Lab.objects.get(name__iexact=lab_name) # Case-insensitive search for existing lab
                    except Lab.DoesNotExist:
                        user, created = User.objects.get_or_create(username=f"lab_{lab_name.replace(' ', '_').lower()}", defaults={'is_active': True})
                        if created:
                            user.set_password('defaultpassword') # IMPORTANT: Change this for production!
                            user.save()
                        lab = Lab.objects.create(
                            user=user,
                            name=lab_name,
                            address=row_data.get('address', ''),
                            city=row_data.get('city', ''),
                            state=row_data.get('state', ''),
                            zip_code=row_data.get('zip code', ''),
                            phone_number=row_data.get('phone number', ''),
                            contact_email=row_data.get('contact email', 'noreply@example.com'),
                            contact_phone=row_data.get('contact phone', '000-000-0000')
                        )
                        messages.info(request, f"Created new lab: {lab_name}")

                    # Create or get Test and associate with Lab
                    test_name = row_data.get('test name')
                    if test_name:
                        # Case-insensitive search for existing test
                        try:
                            test = Test.objects.get(name__iexact=test_name)
                            created = False
                        except Test.DoesNotExist:
                            test = Test.objects.create(
                                name=test_name,
                                description=row_data.get('test description', ''),
                                price=row_data.get('test price', 0.00)
                            )
                            created = True
                        
                        if created:
                            messages.info(request, f"Created new test: {test_name}")
                        lab.tests.add(test)
                        messages.info(request, f"Associated test '{test_name}' with lab '{lab_name}'")

                messages.success(request, "Excel file uploaded and processed successfully!")
                return redirect('admin_lab_list') # Redirect to admin lab list after successful upload

            except Exception as e:
                messages.error(request, f"Error processing Excel file: {e}")
    else:
        form = ExcelUploadForm()
    return render(request, 'admin_upload_excel.html', {'form': form})


@login_required
def lab_upload_tests_excel(request):
    if not hasattr(request.user, 'lab'):
        messages.error(request, "You must be a registered lab to upload tests.")
        return redirect('lab_registration')

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                # Assuming the first row is headers
                headers = [cell.value.lower() if cell.value else None for cell in sheet[1]]
                expected_headers = ['name', 'description', 'price']

                # Fix 1: Map "test name" header to "name" if present
                if 'test name' in headers and 'name' not in headers:
                    headers[headers.index('test name')] = 'name'

                if not all(header in headers for header in expected_headers):
                    messages.error(request, "Missing required columns (case-insensitive). Accepts: Name/Test Name, Description, Price.")
                    return render(request, 'labmanage.html', {'form': form, 'lab': request.user.lab, 'tests': request.user.lab.tests.all()})

                for row_index in range(2, sheet.max_row + 1):
                    row_data = {headers[i]: cell.value for i, cell in enumerate(sheet[row_index])}

                    test_name = row_data.get('name')
                    test_description = row_data.get('description', '')
                    test_price = row_data.get('price', 0.00)

                    if not test_name:
                        messages.warning(request, f"Skipping row {row_index}: Test Name is missing.")
                        continue

                    test, created = Test.objects.get_or_create(
                        name__iexact=test_name, # Case-insensitive search
                        defaults={
                            'description': test_description,
                            'price': test_price
                        }
                    )
                    
                    # Explicitly update fields even if the test already existed
                    test.name = test_name # Ensure name is set/updated
                    test.description = test_description
                    test.price = test_price
                    test.save()

                    if created:
                        messages.info(request, f"Created new test: {test_name}")
                    else:
                        messages.info(request, f"Updated existing test: {test_name}")

                    request.user.lab.tests.add(test)
                    messages.info(request, f"Associated test '{test_name}' with your lab.")

                messages.success(request, "Excel file uploaded and processed successfully!")
                # Fix 2: Redirect to manage_lab to preserve original UI context
                return redirect('manage_lab')

            except Exception as e:
                messages.error(request, f"Error processing Excel file: {e}")
    
    # Get lab and tests
    lab = request.user.lab
    tests = lab.tests.all()
    lab_messages = ContactMessage.objects.filter(lab=lab, recipient_admin=False).order_by('-sent_at')
    
    context = {
        'form': TestForm(),  # Form for "Add a New Test" section
        'lab': lab,
        'tests': tests,
        'excel_upload_form': ExcelUploadForm(),  # Form for Excel upload section
        'lab_messages': lab_messages,
        'total_tests': tests.count(),
        'total_messages': lab_messages.count(),
        'recent_messages': lab_messages[:5],
    }
    return render(request, 'labmanage.html', context)


@login_required
def delete_message(request, message_id):
    message = get_object_or_404(ContactMessage, id=message_id)
    # Ensure only the recipient lab can delete the message
    if message.lab and message.lab.user == request.user:
        message.delete()
        messages.success(request, 'Message deleted successfully.')
    else:
        messages.error(request, 'You are not authorized to delete this message.')
    return redirect('manage_lab')

@login_required
@user_passes_test(lambda user: user.is_staff)
def admin_delete_message(request, message_id):
    message = get_object_or_404(ContactMessage, id=message_id)
    message.delete()
    messages.success(request, 'Message deleted successfully.')
    return redirect('view_contacts')
    # Ensure that only the lab to whom the message was sent can delete it
    if not (hasattr(request.user, 'lab') and message.lab == request.user.lab):
        messages.error(request, "You are not authorized to delete this message.")
        return redirect('manage_lab')

    if request.method == 'POST':
        message.delete()
        messages.success(request, "Message deleted successfully.")
        return redirect('manage_lab')
    
    # Optionally, you could render a confirmation page here
    return redirect('manage_lab') # For now, just redirect back to manage_lab


# Helper function for AI booking processing
def _process_ai_booking(user_message, session_id, user):
    """Collect booking details (name, email, phone) - doesn't create booking yet"""
    import re
    from datetime import datetime, timedelta
    from django.core.cache import cache
    
    # Check if user is providing symptoms/health concerns (not booking details yet)
    symptoms_keywords = ['feel', 'pain', 'tired', 'fatigue', 'weak', 'fever', 'headache', 'worry', 'concern', 'symptom', 'problem', 'issue', 'sick', 'ill', 'disease', 'diabetes', 'heart', 'thyroid', 'liver', 'kidney', 'chest', 'stomach', 'blood pressure']
    
    has_symptoms = any(word in user_message.lower() for word in symptoms_keywords) and '@' not in user_message
    has_email = '@' in user_message
    has_name = any(phrase in user_message.lower() for phrase in ['my name is', 'i am', 'name:'])
    
    # Extract booking details using regex patterns
    name_match = re.search(r'(?:my name is|i am|name:?\s*)\s*([A-Za-z\s]+?)(?:,|email|$)', user_message, re.IGNORECASE)
    email_match = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', user_message)
    phone_match = re.search(r'\b(\d{10}|\d{3}[-.\s]?\d{3}[-.\s]?\d{4})\b', user_message)
    
    # Clean extracted data
    patient_name = name_match.group(1).strip() if name_match else None
    patient_email = email_match.group(1) if email_match else None
    patient_phone = phone_match.group(1) if phone_match else None
    
    # Check for common date patterns
    user_lower = user_message.lower()
    
    # Tomorrow patterns
    if 'tomorrow' in user_lower:
        preferred_date = datetime.now() + timedelta(days=1)
        # Check for specific time mentions (requires : or am/pm)
        time_match = re.search(r'(?:at\s+)?(\d{1,2}):(\d{2})\s*(?:am|pm)?|(?:at\s+)?(\d{1,2})\s*(?:am|pm)\b', user_message, re.IGNORECASE)
        if time_match:
            if time_match.group(1):  # Has colon (HH:MM format)
                hour = int(time_match.group(1))
                minute = int(time_match.group(2))
            else:  # Has AM/PM (H AM/PM format)
                hour = int(time_match.group(3))
                minute = 0
            
            if hour <= 12 and 'pm' in user_message.lower():
                if hour != 12:
                    hour += 12
            elif hour > 12 and 'am' in user_message.lower():
                hour = 0
            
            preferred_date = preferred_date.replace(hour=hour, minute=minute, second=0, microsecond=0)
        elif 'morning' in user_lower:
            preferred_date = preferred_date.replace(hour=9, minute=0, second=0, microsecond=0)
        elif 'afternoon' in user_lower or 'noon' in user_lower:
            preferred_date = preferred_date.replace(hour=14, minute=0, second=0, microsecond=0)
        elif 'evening' in user_lower:
            preferred_date = preferred_date.replace(hour=18, minute=0, second=0, microsecond=0)
        else:
            preferred_date = preferred_date.replace(hour=10, minute=0, second=0, microsecond=0)
    
    # Specific date patterns (e.g., "12 feb", "12 feb 1:00 am")
    date_match = re.search(r'(\d{1,2})\s*(?:feb|january|february|march|april|may|june|july|august|september|oct|october|nov|november|december|jan|mar|apr|jun|jul|aug|sep|sept|dec)\b', user_message, re.IGNORECASE)
    if date_match:
        day = int(date_match.group(1))
        month_str = re.search(r'(?:feb|january|february|march|april|may|june|july|august|september|oct|october|nov|november|december|jan|mar|apr|jun|jul|aug|sep|sept|dec)', user_message, re.IGNORECASE).group(0).lower()
        
        # Month mapping
        months = {
            'jan': 1, 'january': 1,
            'feb': 2, 'february': 2,
            'mar': 3, 'march': 3,
            'apr': 4, 'april': 4,
            'may': 5,
            'jun': 6, 'june': 6,
            'jul': 7, 'july': 7,
            'aug': 8, 'august': 8,
            'sep': 9, 'sept': 9, 'september': 9,
            'oct': 10, 'october': 10,
            'nov': 11, 'november': 11,
            'dec': 12, 'december': 12
        }
        
        month = months.get(month_str, datetime.now().month)
        year = datetime.now().year
        
        try:
            preferred_date = datetime(year=year, month=month, day=day, hour=10, minute=0)
            
            # If date is in the past this year, try next year
            if preferred_date < datetime.now():
                preferred_date = datetime(year=year+1, month=month, day=day, hour=10, minute=0)
            
            # Check for specific time
            time_match = re.search(r'(\d{1,2}):(\d{2})\s*(?:am|pm)?', user_message, re.IGNORECASE)
            if time_match:
                hour = int(time_match.group(1))
                minute = int(time_match.group(2))
                if 'pm' in user_message.lower() and hour < 12:
                    hour += 12
                elif 'am' in user_message.lower() and hour == 12:
                    hour = 0
                preferred_date = preferred_date.replace(hour=hour, minute=minute)
        except ValueError:
            # Invalid date, keep current time
            preferred_date = datetime.now()
    
    # Now handle test name extraction
    test_name = None
    
    # FIRST: Check if one was stored in cache from previous message (HIGHEST PRIORITY)
    stored_test = cache.get(f"booking_test_{session_id}")
    if stored_test:
        test_name = stored_test
    
    # If user provided symptoms + name + email but no explicit test name, recommend tests based on symptoms
    if has_symptoms and has_name and has_email and not test_name:
        return _get_symptom_based_recommendations_for_booking(user_message, patient_name, patient_email, patient_phone, session_id)
    
    # Validate extracted data
    if not patient_name or not patient_email:
        missing = []
        if not patient_name: missing.append("**name**")
        if not patient_email: missing.append("**email**")
        
        message = f"❌ **Missing Information**\n\n"
        message += f"I need your {' and '.join(missing)}.\n\n"
        message += "Please provide in this format:\n"
        message += "*Example:* My name is John Smith, john@gmail.com, 9876543210\n"
        return {'success': False, 'message': message}
    
    # Get the cached test name (after validating name and email)
    test_name = cache.get(f"booking_test_{session_id}")
    if not test_name:
        message = f"❌ **Test Not Selected**\n\n"
        message += "Please first select the test you want to book.\n"
        message += "Example: *Book Blood Sugar*"
        return {'success': False, 'message': message}
    
    # Find matching test from database
    matching_test = Test.objects.filter(name__icontains=test_name).first()
    if not matching_test:
        # Get available tests to suggest
        available_tests = Test.objects.all()[:5]
        if available_tests.exists():
            test_list = ", ".join([f"**{t.name}**" for t in available_tests])
            message = f"❌ **Test Not Found**\n\n"
            message += f"I couldn't find a test called \"{test_name}\".\n\n"
            message += f"**Available tests:** {test_list}\n\n"
            message += "Please mention one of these tests or describe your symptoms."
        else:
            message = f"❌ **No Tests Available**\n\n"
            message += "Currently no tests are available in the system.\n"
            message += "Please contact our support team."
        return {'success': False, 'message': message}
    
    # Find a lab that offers this test
    lab = Lab.objects.filter(tests=matching_test).first()
    if not lab:
        message = f"❌ **Lab Not Found**\n\n"
        message += f"Unfortunately, no lab currently offers the **{matching_test.name}** test.\n"
        message += "Please try describing your symptoms so I can recommend available tests."
        return {'success': False, 'message': message}
    
    # STAGE 1 COMPLETE: Save details to cache and ask for date/time selection
    booking_details = {
        'name': patient_name,
        'email': patient_email,
        'phone': patient_phone,
        'test_name': matching_test.name,
        'test_id': matching_test.id,
        'lab_id': lab.id
    }
    cache.set(f"booking_details_{session_id}", booking_details, 1800)  # 30 min TTL
    
    message = f"✅ **Great! Details Confirmed**\n\n"
    message += f"📋 **Your Information:**\n"
    message += f"• Name: {patient_name}\n"
    message += f"• Email: {patient_email}\n"
    if patient_phone:
        message += f"• Phone: {patient_phone}\n"
    message += f"\n🩸 **Test:** {matching_test.name}\n"
    message += f"💰 **Price:** Rs. {matching_test.price if matching_test.price else 'Contact Lab'}\n"
    message += f"🏥 **Lab:** {lab.name}\n\n"
    message += f"**Now, please select your preferred appointment date and time:**\n\n"
    message += f"• **Today** at 10:00 AM\n"
    message += f"• **Tomorrow** at 9:00 AM\n"
    message += f"• **Tomorrow** at 2:00 PM\n"
    message += f"• **This Week** (specify date)\n"
    message += f"• **Custom date/time**: Just tell me when! (e.g., *14 Feb at 3:30 PM*)\n\n"
    message += f"What time works best for you?"
    
    return {
        'success': False,  # Not booking yet, asking for date/time
        'message': message,
        'suggestions': ["Today Morning", "Tomorrow", "This Week", "Custom Date"]
    }


def _process_date_selection(user_message, session_id, user):
    """Process date/time selection and create the booking"""
    import re
    from datetime import datetime, timedelta
    from django.core.cache import cache
    from .email_utils import send_booking_confirmation_email
    
    # Get saved booking details
    booking_details = cache.get(f"booking_details_{session_id}")
    if not booking_details:
        message = f"❌ **Session Expired**\n\n"
        message += "Your booking session has expired. Please start over:\n"
        message += "*Example: Book Blood Sugar*"
        return {'success': False, 'message': message}
    
    # Extract date/time from user message
    user_lower = user_message.lower()
    preferred_date = datetime.now() + timedelta(days=1)  # Default to tomorrow
    
    # Parse date preferences
    if 'today' in user_lower:
        preferred_date = datetime.now()
        if 'morning' in user_lower or '9' in user_message:
            preferred_date = preferred_date.replace(hour=9, minute=0, second=0, microsecond=0)
        elif 'afternoon' in user_lower or '2' in user_message or '14' in user_message:
            preferred_date = preferred_date.replace(hour=14, minute=0, second=0, microsecond=0)
        else:
            preferred_date = preferred_date.replace(hour=10, minute=0, second=0, microsecond=0)
    
    elif 'tomorrow' in user_lower:
        preferred_date = datetime.now() + timedelta(days=1)
        # Check for specific time
        if '9' in user_message or 'morning' in user_lower:
            preferred_date = preferred_date.replace(hour=9, minute=0, second=0, microsecond=0)
        elif '2' in user_message or '14' in user_message or 'afternoon' in user_lower:
            preferred_date = preferred_date.replace(hour=14, minute=0, second=0, microsecond=0)
        else:
            preferred_date = preferred_date.replace(hour=10, minute=0, second=0, microsecond=0)
    
    else:
        # Check for specific date patterns (e.g., "12 feb", "14 feb 3:30 pm")
        date_match = re.search(r'(\d{1,2})\s*(?:feb|january|february|march|april|may|june|july|august|september|oct|october|nov|november|december|jan|mar|apr|jun|jul|aug|sep|sept|dec)\b', user_message, re.IGNORECASE)
        if date_match:
            day = int(date_match.group(1))
            month_str = re.search(r'(?:feb|january|february|march|april|may|june|july|august|september|oct|october|nov|november|december|jan|mar|apr|jun|jul|aug|sep|sept|dec)', user_message, re.IGNORECASE).group(0).lower()
            
            months = {
                'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
                'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
                'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9, 'oct': 10, 'october': 10,
                'nov': 11, 'november': 11, 'dec': 12, 'december': 12
            }
            
            month = months.get(month_str, datetime.now().month)
            year = datetime.now().year
            
            try:
                preferred_date = datetime(year=year, month=month, day=day, hour=10, minute=0)
                if preferred_date < datetime.now():
                    preferred_date = datetime(year=year+1, month=month, day=day, hour=10, minute=0)
                
                # Check for time mention
                time_match = re.search(r'(\d{1,2}):(\d{2})\s*(?:am|pm)?', user_message, re.IGNORECASE)
                if time_match:
                    hour = int(time_match.group(1))
                    minute = int(time_match.group(2))
                    if 'pm' in user_message.lower() and hour < 12:
                        hour += 12
                    elif 'am' in user_message.lower() and hour == 12:
                        hour = 0
                    preferred_date = preferred_date.replace(hour=hour, minute=minute)
            except ValueError:
                preferred_date = datetime.now() + timedelta(days=1)
    
    # Create the booking
    try:
        booking = TestBooking.objects.create(
            name=booking_details['name'],
            email=booking_details['email'],
            test_id=booking_details['test_id'],
            lab_id=booking_details['lab_id'],
            booking_date=preferred_date,
            status='booked'
        )
        
        # Send confirmation email
        try:
            send_booking_confirmation_email(booking)
        except Exception as e:
            print(f"Email sending error: {str(e)}")
        
        # Get test and lab for response
        test = booking.test
        lab = booking.lab
        
        # Success response
        message = f"✅ **Booking Confirmed!**\n\n"
        message += f"Great news, {booking_details['name']}! Your test booking is ready.\n\n"
        message += f"**Booking Details:**\n"
        message += f"🎫 **Booking ID:** {booking.booking_id}\n"
        message += f"🩸 **Test:** {test.name}\n"
        message += f"💰 **Price:** Rs. {test.price if test.price else 'Contact Lab'}\n"
        message += f"🏥 **Lab:** {lab.name}\n"
        message += f"📍 **Location:** {lab.city}, {lab.state}\n"
        message += f"📅 **Appointment:** {preferred_date.strftime('%B %d, %Y at %I:%M %p')}\n"
        message += f"✓ **Status:** Booked\n\n"
        message += f"📧 Confirmation sent to **{booking_details['email']}**\n"
        if booking_details['phone']:
            message += f"📱 Phone: {booking_details['phone']}\n"
        message += f"📞 Lab Contact: {lab.contact_phone}\n"
        message += f"💬 Lab Email: {lab.contact_email}"
        
        return {'success': True, 'message': message}
    
    except Exception as e:
        message = f"❌ **Booking Failed**\n\n"
        message += f"An error occurred: {str(e)}\n"
        message += "Please try again or contact support."
        return {'success': False, 'message': message}


def _get_symptom_based_recommendations_for_booking(user_message, patient_name, patient_email, patient_phone, session_id):
    """Get test recommendations based on symptoms and auto-book if possible"""
    import re
    from datetime import datetime, timedelta
    from .email_utils import send_booking_confirmation_email
    from lab_suggestion.ai_service import AIRecommendationService
    
    # Extract preferred appointment date/time from message
    preferred_date = datetime.now()
    
    # Check for common date patterns
    user_lower = user_message.lower()
    
    # Tomorrow patterns
    if 'tomorrow' in user_lower:
        preferred_date = datetime.now() + timedelta(days=1)
        # Check for specific time mentions (requires : or am/pm)
        time_match = re.search(r'(?:at\s+)?(\d{1,2}):(\d{2})\s*(?:am|pm)?|(?:at\s+)?(\d{1,2})\s*(?:am|pm)\b', user_message, re.IGNORECASE)
        if time_match:
            if time_match.group(1):  # Has colon (HH:MM format)
                hour = int(time_match.group(1))
                minute = int(time_match.group(2))
            else:  # Has AM/PM (H AM/PM format)
                hour = int(time_match.group(3))
                minute = 0
            
            if hour <= 12 and 'pm' in user_message.lower():
                if hour != 12:
                    hour += 12
            elif hour > 12 and 'am' in user_message.lower():
                hour = 0
            
            preferred_date = preferred_date.replace(hour=hour, minute=minute, second=0, microsecond=0)
        elif 'morning' in user_lower:
            preferred_date = preferred_date.replace(hour=9, minute=0, second=0, microsecond=0)
        else:
            preferred_date = preferred_date.replace(hour=10, minute=0, second=0, microsecond=0)
    
    # Specific date patterns
    date_match = re.search(r'(\d{1,2})\s*(?:feb|january|february|march|april|may|june|july|august|september|oct|october|nov|november|december|jan|mar|apr|jun|jul|aug|sep|sept|dec)\b', user_message, re.IGNORECASE)
    if date_match:
        day = int(date_match.group(1))
        month_str = re.search(r'(?:feb|january|february|march|april|may|june|july|august|september|oct|october|nov|november|december|jan|mar|apr|jun|jul|aug|sep|sept|dec)', user_message, re.IGNORECASE).group(0).lower()
        
        months = {
            'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
            'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
            'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9, 'oct': 10, 'october': 10,
            'nov': 11, 'november': 11, 'dec': 12, 'december': 12
        }
        
        month = months.get(month_str, datetime.now().month)
        year = datetime.now().year
        
        try:
            preferred_date = datetime(year=year, month=month, day=day, hour=10, minute=0)
            if preferred_date < datetime.now():
                preferred_date = datetime(year=year+1, month=month, day=day, hour=10, minute=0)
            
            time_match = re.search(r'(\d{1,2}):(\d{2})\s*(?:am|pm)?', user_message, re.IGNORECASE)
            if time_match:
                hour = int(time_match.group(1))
                minute = int(time_match.group(2))
                if 'pm' in user_message.lower() and hour < 12:
                    hour += 12
                elif 'am' in user_message.lower() and hour == 12:
                    hour = 0
                preferred_date = preferred_date.replace(hour=hour, minute=minute)
        except ValueError:
            preferred_date = datetime.now()
    
    # Get recommendations based on symptoms
    recommendation_service = AIRecommendationService()
    recommendation = recommendation_service.recommend_tests(user_message)
    
    if not recommendation or not recommendation.recommended_tests.exists():
        return {
            'success': False,
            'message': '❌ Could not find matching tests for your symptoms.\n\nPlease specify which test you want to book or contact us for help.'
        }
    
    # Get the first/best recommended test
    recommended_tests = recommendation.recommended_tests.all()[:5]
    
    if recommended_tests.count() == 1:
        # Auto-book the single recommended test
        test = recommended_tests.first()
        lab = Lab.objects.filter(tests=test).first()
        
        if not lab:
            return {
                'success': False,
                'message': f"❌ No lab offers this test: {test.name}\n\nTry another test or symptom."
            }
        
        try:
            booking = TestBooking.objects.create(
                name=patient_name,
                email=patient_email,
                test=test,
                lab=lab,
                booking_date=preferred_date,  # Use extracted preferred date/time
                status='booked'
            )
            
            try:
                send_booking_confirmation_email(booking)
            except Exception as e:
                print(f"Email sending error: {str(e)}")
            
            message = f"✅ **Booking Confirmed!**\n\n"
            message += f"Perfect! Based on your symptoms, I've booked the ideal test.\n\n"
            message += f"**Booking Details:**\n"
            message += f"🎫 **Booking ID:** {booking.booking_id}\n"
            message += f"🩸 **Test:** {test.name}\n"
            message += f"💰 **Price:** Rs. {test.price if test.price else 'Contact Lab'}\n"
            message += f"🏥 **Lab:** {lab.name}\n"
            message += f"📍 **Location:** {lab.city}, {lab.state}\n"
            message += f"📅 **Booked Date:** {booking.booking_date.strftime('%B %d, %Y at %I:%M %p')}\n\n"
            message += f"📧 Confirmation sent to **{patient_email}**\n"
            if patient_phone:
                message += f"📱 Phone: {patient_phone}\n"
            message += f"📞 Lab Contact: {lab.contact_phone}"
            
            return {'success': True, 'message': message}
        except Exception as e:
            return {'success': False, 'message': f"❌ Booking failed: {str(e)}"}
    else:
        # Multiple recommendations - ask user to choose
        test_list = "\n".join([f"• **{t.name}** - Rs. {t.price if t.price else 'Contact Lab'}" for t in recommended_tests])
        message = f"✨ **Perfect Match: {recommended_tests.count()} Tests Found!**\n\n"
        message += f"Based on your symptoms of '{user_message}', here are the recommended tests:\n\n"
        message += test_list
        message += f"\n\nWhich one would you like to book? Just reply with the test name!\n"
        message += f"Example: *I want to book BLOOD SUGAR F*"
        
        return {'success': False, 'message': message}


def _get_symptom_based_recommendations(user_message):
    """Get test recommendations based on user symptoms"""
    from .models import Test, Lab
    
    user_lower = user_message.lower()
    
    # Symptom to test mapping (only use tests that exist in database)
    all_tests = Test.objects.all()
    test_dict = {test.name.lower(): test for test in all_tests}
    
    recommendations = []
    
    # Map symptoms to test keywords
    if any(word in user_lower for word in ['diabetes', 'blood sugar', 'glucose', 'thirsty', 'urination', 'sweet']):
        keywords = ['glucose', 'blood', 'cbc', 'fasting']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    elif any(word in user_lower for word in ['heart', 'chest', 'cardiac', 'breathless', 'angina']):
        keywords = ['cardiac', 'ecg', 'heart', 'troponin']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    elif any(word in user_lower for word in ['thyroid', 'tired', 'fatigue', 'weight', 'tsh']):
        keywords = ['thyroid', 'tsh', 't3', 't4']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    elif any(word in user_lower for word in ['liver', 'jaundice', 'yellow', 'liver function']):
        keywords = ['liver', 'alt', 'ast', 'bilirubin']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    elif any(word in user_lower for word in ['kidney', 'kidney function', 'creatinine']):
        keywords = ['kidney', 'creatinine', 'bun']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    elif any(word in user_lower for word in ['fever', 'infection', 'cold', 'cough', 'wbc']):
        keywords = ['cbc', 'blood', 'infection', 'wbc']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    elif any(word in user_lower for word in ['cholesterol', 'lipid', 'fat', 'triglyceride']):
        keywords = ['lipid', 'cholesterol', 'triglyceride']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    else:
        # General health check
        keywords = ['blood', 'cbc', 'general']
        for test in all_tests:
            if any(kw in test.name.lower() for kw in keywords):
                recommendations.append(test)
    
    # Remove duplicates while preserving order
    seen = set()
    unique_recommendations = []
    for test in recommendations:
        if test.id not in seen:
            seen.add(test.id)
            unique_recommendations.append(test)
    
    if unique_recommendations:
        message = f"👨‍⚕️ **Test Recommendations Based on Your Symptoms**\n\n"
        message += f"Based on what you shared, here are the tests I recommend:\n\n"
        
        for i, test in enumerate(unique_recommendations[:5], 1):
            labs_offering = Lab.objects.filter(tests=test).count()
            message += f"{i}. **{test.name}**\n"
            if test.description:
                message += f"   *{test.description}*\n"
            if test.price:
                message += f"   💰 Price: Rs. {test.price}\n"
            if labs_offering > 0:
                message += f"   🏥 Available at {labs_offering} lab(s)\n"
            message += "\n"
        
        message += "**Next Steps:**\n"
        message += "• To book one of these tests, just say: *I want to book [test name], my name is [your name], [your email]*\n"
        message += "• Or ask me more questions about any of these tests\n"
        message += "• I can also help you find the best lab for your test"
        
        return {'success': False, 'message': message}
    else:
        message = f"🤔 **No Specific Tests Found**\n\n"
        message += "I couldn't find specific tests for your symptoms in the current database.\n\n"
        message += "**Available tests:**\n"
        for test in Test.objects.all()[:10]:
            message += f"• {test.name}\n"
        
        message += "\nPlease tell me which test you'd like to book with your details!"
        
        return {'success': False, 'message': message}


# AI Chatbot Views
# Note: Full-page chatbot removed - using floating widget only (more popular pattern)
# The floating widget is available on all pages via base.html

@csrf_exempt
@require_http_methods(["POST"])
def chatbot_api(request):
    """API endpoint for chatbot interactions"""
    try:
        from django.core.cache import cache
        data = json.loads(request.body)
        user_message = data.get('message', '').strip()
        session_id = data.get('session_id', str(uuid.uuid4()))
        
        if not user_message:
            return JsonResponse({'error': 'Message is required'}, status=400)
        
        # Initialize chatbot service
        chatbot = AIChatbotService()
        
        # Check if user is trying to book or provided booking details
        user_message_lower = user_message.lower()
        is_booking_intent = any(word in user_message_lower for word in ['book', 'reserve', 'schedule', 'appointment'])
        is_booking_details = any(word in user_message_lower for word in ['my name is', 'i am', 'email:', 'email is', 'book for me', '@'])
        
        # First, check if this message contains a test name and store it in cache
        from .models import Test
        all_tests = Test.objects.all()
        detected_test = None
        
        import re
        # Look for test keywords in the message
        for test in all_tests:
            if test.name.lower() in user_message_lower:
                detected_test = test.name
                break
        
        # If no exact match, try regex extraction for "book [test name]" pattern
        if not detected_test:
            test_match = re.search(r'(?:book|test:?\s*)\s*([A-Za-z\s0-9/:-]+?)(?:,|email|my|$)', user_message, re.IGNORECASE)
            if test_match:
                test_name_candidate = test_match.group(1).strip().lower()
                # Fuzzy match - find test that contains the candidate keywords
                for test in all_tests:
                    test_name_lower = test.name.lower()
                    # Check if test name contains most keywords
                    if test_name_lower.startswith(test_name_candidate) or test_name_candidate in test_name_lower:
                        detected_test = test.name
                        break
        
        # Store detected test in cache for later use (15 minute expiry)
        if detected_test:
            cache.set(f"booking_test_{session_id}", detected_test, 900)
        
        # Check if this is a booking attempt (has details like name/email)
        current_booking_stage = cache.get(f"booking_stage_{session_id}")
        
        # If user is in the middle of a booking session, check stage first
        if current_booking_stage == 'date_selection':
            # User is selecting date/time (don't check for name/email requirement)
            booking_result = _process_date_selection(user_message, session_id, request.user)
            if booking_result['success']:
                # Booking complete!
                bot_response = booking_result['message']
                suggestions = ["View my bookings", "Book another test", "Go to home"]
                cache.delete(f"booking_stage_{session_id}")
                cache.delete(f"booking_test_{session_id}")
                cache.delete(f"booking_details_{session_id}")
            else:
                # Ask to select date/time again
                bot_response = booking_result['message']
                suggestions = booking_result.get('suggestions', ["Today", "Tomorrow", "This Week"])
        
        elif is_booking_details and (is_booking_intent or detected_test or cache.get(f"booking_test_{session_id}")):
            # Stage 1: Collect name, email, phone
            booking_result = _process_ai_booking(user_message, session_id, request.user)
            bot_response = booking_result['message']
            suggestions = booking_result.get('suggestions', ["Today", "Tomorrow", "This Week"])
            
            # Check if we have the booking details (ask for date/time next)
            if 'Details Confirmed' in bot_response or 'Now, please select' in bot_response:
                # Move to date selection stage
                cache.set(f"booking_stage_{session_id}", 'date_selection', 1800)  # 30 min TTL
            else:
                suggestions = ["Try again with correct details", "What tests do you have?", "Find labs near me"]
        else:
            # Generate normal response
            bot_response, suggestions = chatbot.generate_response(user_message, session_id)
        
        # Save to database
        chat_message = ChatMessage.objects.create(
            session_id=session_id,
            user_message=user_message,
            bot_response=bot_response,
            user=request.user if request.user.is_authenticated else None
        )
        
        return JsonResponse({
            'response': bot_response,
            'suggestions': suggestions,
            'session_id': session_id,
            'timestamp': chat_message.created_at.isoformat()
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def ai_recommendations_view(request):
    """View for AI-powered test recommendations based on symptoms"""
    recommended_tests = None
    symptoms = ''
    recommendation = None
    
    if request.method == 'POST':
        symptoms = request.POST.get('symptoms', '').strip()
        
        if symptoms:
            recommendation_service = AIRecommendationService()
            recommendation = recommendation_service.recommend_tests(
                symptoms,
                user=request.user if request.user.is_authenticated else None
            )
            
            if recommendation:
                recommended_tests = recommendation.recommended_tests.all()
                if not recommended_tests.exists():
                    messages.warning(request, 'Recommendations were created but no tests were found. Please ensure sample data is loaded.')
            else:
                messages.info(request, 'No specific recommendations found. Please try describing your symptoms in more detail.')
        else:
            messages.error(request, 'Please enter your symptoms.')
    
    return render(request, 'ai_recommendations.html', {
        'recommendation': recommendation,
        'recommended_tests': recommended_tests,
        'symptoms': symptoms
    })


def chatbot_history(request):
    """View chat history for authenticated users"""
    if not request.user.is_authenticated:
        messages.info(request, 'Please login to view your chat history.')
        return redirect('login')
    
    chat_history = ChatMessage.objects.filter(user=request.user).order_by('-created_at')[:50]
    return render(request, 'chatbot_history.html', {'chat_history': chat_history})


# ======================== TEST BOOKING VIEWS ========================

def book_test(request, test_id, lab_id):
    """View for booking a test"""
    test = get_object_or_404(Test, id=test_id)
    lab = get_object_or_404(Lab, id=lab_id)
    
    if request.method == 'POST':
        form = TestBookingForm(request.POST)
        if form.is_valid():
            booking = form.save(commit=False)
            booking.test = test
            booking.lab = lab
            booking.save()
            
            # Send confirmation email
            try:
                email_sent = send_booking_confirmation_email(booking)
                if email_sent:
                    messages.success(request, f'Your Booking has been Booked! A confirmation email has been sent to {booking.email}')
                else:
                    messages.warning(request, f'Your Booking has been Booked! However, we could not send the confirmation email.')
            except Exception as e:
                print(f"Email sending error: {str(e)}")
                messages.warning(request, f'Your Booking has been Booked! However, there was an issue sending the confirmation email.')
            
            return render(request, 'booking_confirmation.html', {
                'booking': booking,
                'test': test,
                'lab': lab
            })
    else:
        form = TestBookingForm()
    
    return render(request, 'book_test.html', {
        'form': form,
        'test': test,
        'lab': lab
    })


def check_booking_status(request):
    """View to check booking status"""
    booking = None
    error_message = None
    
    if request.method == 'POST':
        booking_id = request.POST.get('booking_id', '').strip()
        email = request.POST.get('email', '').strip()
        
        if booking_id and email:
            try:
                booking = TestBooking.objects.get(booking_id=booking_id, email=email)
            except TestBooking.DoesNotExist:
                error_message = 'No booking found with this ID and email. Please check and try again.'
        else:
            error_message = 'Please provide both Booking ID and Email.'
    
    return render(request, 'check_booking_status.html', {
        'booking': booking,
        'error_message': error_message
    })


def update_booking(request, booking_id):
    """View to update booking date or details"""
    booking = get_object_or_404(TestBooking, booking_id=booking_id, status='booked')
    
    if request.method == 'POST':
        form = TestBookingForm(request.POST, instance=booking)
        # Verify email matches
        email = request.POST.get('email', '').strip()
        if email != booking.email:
            messages.error(request, 'Email does not match the booking email.')
            return render(request, 'update_booking.html', {'form': form, 'booking': booking})
        
        if form.is_valid():
            form.save()
            
            # Send update notification email
            try:
                email_sent = send_booking_update_email(booking)
                if email_sent:
                    messages.success(request, f'Your booking has been updated successfully! A confirmation email has been sent.')
                else:
                    messages.warning(request, f'Your booking has been updated successfully! However, we could not send the notification email.')
            except Exception as e:
                print(f"Email sending error: {str(e)}")
                messages.warning(request, f'Your booking has been updated successfully! However, there was an issue sending the notification email.')
            
            return redirect('check_booking_status')
    else:
        form = TestBookingForm(instance=booking)
    
    return render(request, 'update_booking.html', {
        'form': form,
        'booking': booking
    })


def cancel_booking(request, booking_id):
    """View to cancel a booking"""
    booking = get_object_or_404(TestBooking, booking_id=booking_id, status='booked')
    
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        # Verify email matches
        if email != booking.email:
            messages.error(request, 'Email does not match the booking email.')
            return render(request, 'cancel_booking.html', {'booking': booking})
        
        booking.status = 'cancelled'
        booking.save()
        
        # Send cancellation notification email
        try:
            email_sent = send_booking_cancellation_email(booking)
            if email_sent:
                messages.success(request, f'Your booking {booking.booking_id} has been cancelled successfully. A confirmation email has been sent.')
            else:
                messages.warning(request, f'Your booking {booking.booking_id} has been cancelled successfully. However, we could not send the notification email.')
        except Exception as e:
            print(f"Email sending error: {str(e)}")
            messages.warning(request, f'Your booking {booking.booking_id} has been cancelled successfully. However, there was an issue sending the notification email.')
        
        return redirect('check_booking_status')
    
    return render(request, 'cancel_booking.html', {'booking': booking})


@login_required
def view_lab_bookings(request):
    """Lab can view and manage all bookings for their lab"""
    try:
        lab = request.user.lab
    except Lab.DoesNotExist:
        messages.error(request, 'Lab not found.')
        return redirect('manage_lab')
    
    # Get all bookings for this lab
    all_bookings = TestBooking.objects.filter(lab=lab).order_by('-booking_date')
    
    # Get filter date from request (default to today)
    from datetime import datetime, timedelta
    filter_date_str = request.GET.get('filter_date')
    
    if filter_date_str:
        try:
            filter_date = datetime.strptime(filter_date_str, '%Y-%m-%d').date()
        except ValueError:
            filter_date = datetime.now().date()
    else:
        filter_date = datetime.now().date()
    
    # Get time sort and status filter from request
    time_sort = request.GET.get('time_sort', 'asc')  # 'asc' for 0:00am-12:00pm, 'desc' for 12:00pm-11:59pm
    filter_status = request.GET.get('filter_status', '')  # Empty string means show all
    
    # Filter bookings by date (compare just the date part)
    bookings = [b for b in all_bookings if b.booking_date.date() == filter_date]
    
    # Filter by status if selected
    if filter_status and filter_status in dict(TestBooking.BOOKING_STATUS_CHOICES):
        bookings = [b for b in bookings if b.status == filter_status]
    
    # Sort by time
    if time_sort == 'asc':
        bookings = sorted(bookings, key=lambda x: x.booking_date.time())
    else:  # desc
        bookings = sorted(bookings, key=lambda x: x.booking_date.time(), reverse=True)
    
    # Handle status update
    if request.method == 'POST':
        booking_id = request.POST.get('booking_id')
        new_status = request.POST.get('status')
        
        try:
            booking = TestBooking.objects.get(booking_id=booking_id, lab=lab)
            if new_status in dict(TestBooking.BOOKING_STATUS_CHOICES):
                booking.status = new_status
                booking.save()
                
                # Send update email to user if status changed
                if new_status in ['test_done', 'not_arrived', 'cancelled']:
                    try:
                        send_booking_update_email(booking)
                    except Exception as e:
                        print(f"Email sending error: {str(e)}")
                
                messages.success(request, f'Booking {booking_id} status updated to {new_status}.')
            else:
                messages.error(request, 'Invalid status.')
        except TestBooking.DoesNotExist:
            messages.error(request, 'Booking not found.')
        
        # Redirect back to bookings page with same filter using URL path
        from django.urls import reverse
        return redirect(f"{reverse('view_lab_bookings')}?filter_date={filter_date_str}&time_sort={time_sort}&filter_status={filter_status}")
    
    context = {
        'lab': lab,
        'bookings': bookings,
        'filter_date': filter_date,
        'filter_date_str': filter_date_str or datetime.now().strftime('%Y-%m-%d'),
        'status_choices': TestBooking.BOOKING_STATUS_CHOICES,
        'time_sort': time_sort,
        'filter_status': filter_status,
    }
    return render(request, 'view_lab_bookings.html', context)

